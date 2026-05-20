# mapper.py
import csv
import io
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import (
    INPUT_FILE,
    OUTPUT_EXCEL,
    CSV_COLUMNS,
    CSV_DELIMITER,
    CSV_ENCODINGS,
    PHONE_FIELDS,
    EMAIL_FIELDS,
    MAX_DUPLICATE_CHECK,
)
from validators import (
    clean_phone,
    is_valid_phone,
    is_valid_email,
    parse_date,
    parse_float,
)

REQUIRED_HEADERS = (
    CSV_COLUMNS["id"],
    CSV_COLUMNS["lead_name"],
    CSV_COLUMNS["mobile_phone"],
    CSV_COLUMNS["work_email"],
)

PROGRESS_EVERY = 500

_COMPANY_LEGAL = re.compile(
    r"\b(ооо|ип|зао)\b",
    re.IGNORECASE,
)

# Порядок колонок как в шаблоне импорта Битрикс24
BITRIX_EXCEL_HEADERS = [
    "ID",
    "Название лида",
    "Обращение",
    "Имя",
    "Фамилия",
    "Отчество",
    "Имя, Фамилия",
    "Дата рождения",
    "Адрес",
    "Улица, номер дома",
    "Квартира, офис, комната, этаж",
    "Населенный пункт",
    "Район",
    "Регион",
    "Почтовый индекс",
    "Страна",
    "Рабочий телефон",
    "Мобильный телефон",
    "Номер факса",
    "Домашний телефон",
    "Номер пейджера",
    "Телефон для рассылок",
    "Другой телефон",
    "Корпоративный сайт",
    "Личная страница",
    "Страница Facebook",
    "Страница ВКонтакте",
    "Страница LiveJournal",
    "Микроблог Twitter",
    "Другой сайт",
    "Рабочий e-mail",
    "Частный e-mail",
    "E-mail для рассылок",
    "Другой e-mail",
    "Контакт Facebook",
    "Контакт Telegram",
    "Контакт ВКонтакте",
    "Контакт Viber",
    "Комментарии Instagram",
    "Контакт Битрикс24 Network",
    "Онлайн-чат",
    "Контакт Открытая линия",
    "Другой контакт",
    "Связанный пользователь",
    "Название компании",
    "Должность",
    "Комментарий",
    "Стадия",
    "Дополнительно о стадии",
    "Товар",
    "Цена",
    "Количество",
    "Возможная сумма",
    "Валюта",
    "Источник",
    "Дополнительно об источнике",
    "Доступен для всех",
    "Ответственный",
    "Тип услуги",
    "Новый файл",
    "Доп файл",
    "Источник телефона",
    "Причина отказа",
    "дело",
]

PHONE_HEADERS = frozenset(
    {
        "Рабочий телефон",
        "Мобильный телефон",
        "Номер факса",
        "Домашний телефон",
        "Номер пейджера",
        "Телефон для рассылок",
        "Другой телефон",
    }
)

EMAIL_HEADERS = frozenset(
    {
        "Рабочий e-mail",
        "Частный e-mail",
        "E-mail для рассылок",
        "Другой e-mail",
    }
)

NUM_HEADERS = frozenset({"Цена", "Количество", "Возможная сумма"})

RED_FILL = PatternFill(
    start_color="FFFFC7CE",
    end_color="FFFFC7CE",
    fill_type="solid",
)


def _cell(raw: dict, key: str) -> str:
    v = raw.get(CSV_COLUMNS[key], "")
    return v.strip() if isinstance(v, str) else str(v or "").strip()


def normalize_company(name: str) -> str:
    if not name or not str(name).strip():
        return ""
    s = str(name).strip().lower()
    s = s.replace('"', "").replace("'", "")
    s = _COMPANY_LEGAL.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _raw_get(raw: dict, header: str) -> str:
    v = raw.get(header, "")
    return v.strip() if isinstance(v, str) else str(v or "").strip()


def _iso_to_ddmmyyyy(iso: str) -> str:
    if not iso:
        return ""
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return ""


def _format_ru_decimal(n: float) -> str:
    return f"{n:.2f}".replace(".", ",")


def _format_excel_cell(header: str, raw: dict) -> str:
    if header == "Название компании":
        val = _raw_get(raw, "Название компании") or _raw_get(raw, "Компания")
    else:
        val = _raw_get(raw, header)
    if not val:
        return ""
    if header in PHONE_HEADERS:
        return clean_phone(val)
    if header in EMAIL_HEADERS:
        return val.replace(" ", "").lower()
    if header == "Дата рождения":
        return _iso_to_ddmmyyyy(parse_date(val))
    if header in NUM_HEADERS:
        parsed = parse_float(val)
        st = val.strip()
        if parsed == 0.0 and st and st != "0":
            return val
        return _format_ru_decimal(parsed)
    return val


def _build_excel_row(raw: dict) -> list[str]:
    return [_format_excel_cell(h, raw) for h in BITRIX_EXCEL_HEADERS]


def read_csv(file_path: str) -> list[dict]:
    """Читает CSV. Кодировка и разделитель из config; заголовки с trim."""
    text = None
    for enc in CSV_ENCODINGS:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                text = f.read()
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(
            f"Не удалось прочитать CSV ({file_path}). "
            f"Испробованы кодировки: {', '.join(CSV_ENCODINGS)}"
        )

    buf = io.StringIO(text)
    reader = csv.DictReader(buf, delimiter=CSV_DELIMITER)
    raw_fields = list(reader.fieldnames or [])
    stripped_set = {(h or "").strip() for h in raw_fields}
    missing = [h for h in REQUIRED_HEADERS if h not in stripped_set]
    if missing:
        raise ValueError(
            "В CSV отсутствуют обязательные колонки: "
            + ", ".join(missing)
        )

    leads = []
    for row in reader:
        new_row = {(k or "").strip(): v for k, v in row.items()}
        leads.append(new_row)
    return leads


def map_lead(raw: dict) -> dict:
    """Мапит одну строку CSV в структурированный словарь"""
    birth_raw = _cell(raw, "birth_date")

    _co = raw.get("Компания", "")
    _co = _co.strip() if isinstance(_co, str) else str(_co or "").strip()
    company_name = _co or _cell(raw, "company_name")

    mapped = {
        "id": _cell(raw, "id"),
        "lead_name": _cell(raw, "lead_name"),
        "full_name": _cell(raw, "full_name"),
        "first_name": _cell(raw, "first_name"),
        "last_name": _cell(raw, "last_name"),
        "middle_name": _cell(raw, "middle_name"),
        "salutation": _cell(raw, "salutation"),
        "birth_date_raw": birth_raw,
        "birth_date": parse_date(birth_raw),
        "company_name": company_name,
        "position": _cell(raw, "position"),
        "comment": _cell(raw, "comment"),
        "stage": _cell(raw, "stage"),
        "stage_extra": _cell(raw, "stage_extra"),
        "source": _cell(raw, "source"),
        "source_extra": _cell(raw, "source_extra"),
        "responsible": _cell(raw, "responsible"),
        "service_type": _cell(raw, "service_type"),
        "rejection_reason": _cell(raw, "rejection_reason"),
        "available_for_all": _cell(raw, "available_for_all"),
        "linked_user": _cell(raw, "linked_user"),
        "deal": _cell(raw, "deal"),
    }

    mapped["address"] = {
        "full": _cell(raw, "address"),
        "street": _cell(raw, "street"),
        "apartment": _cell(raw, "apartment"),
        "city": _cell(raw, "city"),
        "district": _cell(raw, "district"),
        "region": _cell(raw, "region"),
        "postal_code": _cell(raw, "postal_code"),
        "country": _cell(raw, "country"),
    }

    phones = []
    for field in PHONE_FIELDS:
        raw_phone = _cell(raw, field)
        if raw_phone:
            cleaned = clean_phone(raw_phone)
            if cleaned and cleaned not in phones:
                phones.append(cleaned)
    mapped["phones"] = phones
    mapped["primary_phone"] = next(
        (p for p in phones if is_valid_phone(p)),
        phones[0] if phones else "",
    )

    emails = []
    for field in EMAIL_FIELDS:
        raw_email = _cell(raw, field)
        if raw_email:
            cleaned = raw_email.replace(" ", "").lower()
            if cleaned and cleaned not in emails:
                emails.append(cleaned)
    mapped["emails"] = emails
    mapped["primary_email"] = next(
        (e for e in emails if is_valid_email(e)),
        emails[0] if emails else "",
    )

    mapped["messengers"] = {
        "telegram": _cell(raw, "contact_telegram"),
        "viber": _cell(raw, "contact_viber"),
        "whatsapp": _cell(raw, "other_contact"),
    }
    mapped["social"] = {
        "vk": _cell(raw, "vk"),
        "facebook": _cell(raw, "facebook"),
        "twitter": _cell(raw, "twitter"),
        "livejournal": _cell(raw, "livejournal"),
        "instagram": _cell(raw, "contact_instagram"),
    }
    mapped["sites"] = {
        "corporate": _cell(raw, "website"),
        "personal": _cell(raw, "personal_page"),
        "other": _cell(raw, "other_site"),
    }

    mapped["product"] = {
        "name": _cell(raw, "product"),
        "price": parse_float(_cell(raw, "price")),
        "quantity": parse_float(_cell(raw, "quantity")),
        "possible_amount": parse_float(_cell(raw, "possible_amount")),
        "currency": _cell(raw, "currency"),
    }

    mapped["files"] = {
        "new_file": _cell(raw, "new_file"),
        "extra_file": _cell(raw, "extra_file"),
    }

    mapped["extra"] = {
        "fax": _cell(raw, "fax"),
        "pager": _cell(raw, "pager"),
        "phone_source": _cell(raw, "phone_source"),
        "contact_bitrix24": _cell(raw, "contact_bitrix24"),
        "contact_online_chat": _cell(raw, "contact_online_chat"),
        "contact_open_line": _cell(raw, "contact_open_line"),
    }

    return mapped


def _build_duplicates(mapped_leads: list[dict]) -> dict:
    phone_to_ids: dict[str, set[str]] = {}
    email_to_ids: dict[str, set[str]] = {}
    company_to_ids: dict[str, set[str]] = {}

    for lead in mapped_leads:
        lid = (lead.get("id") or "").strip()
        if not lid:
            lid = f"row_{lead.get('_csv_row', 0)}"
        for p in lead.get("phones") or []:
            phone_to_ids.setdefault(p, set()).add(lid)
        for e in lead.get("emails") or []:
            email_to_ids.setdefault(e.lower(), set()).add(lid)
        cn = lead.get("company_name") or ""
        key = normalize_company(cn)
        if key:
            company_to_ids.setdefault(key, set()).add(lid)

    def only_multi(d: dict[str, set[str]]) -> dict[str, list[str]]:
        out = {}
        for k, ids in d.items():
            if len(ids) > 1:
                out[k] = sorted(ids)
        return out

    return {
        "phones": only_multi(phone_to_ids),
        "emails": only_multi(email_to_ids),
        "companies": only_multi(company_to_ids),
    }


def _duplicate_groups_count(dup: dict) -> int:
    return (
        len(dup.get("phones") or {})
        + len(dup.get("emails") or {})
        + len(dup.get("companies") or {})
    )


def _lead_ids_in_duplicates(duplicates: dict) -> set[str]:
    ids: set[str] = set()
    for key in ("phones", "emails", "companies"):
        block = duplicates.get(key) or {}
        if not isinstance(block, dict):
            continue
        for _k, id_list in block.items():
            ids.update(id_list)
    return ids


def _error_row_values(raw: dict) -> list[str]:
    row = []
    for h in BITRIX_EXCEL_HEADERS:
        if h == "Название компании":
            row.append(
                _raw_get(raw, "Название компании") or _raw_get(raw, "Компания")
            )
        else:
            row.append(_raw_get(raw, h))
    return row


def _apply_sheet_view(
    ws,
    headers: list[str],
    n_data_rows: int,
    header_row: int = 1,
) -> None:
    ncols = len(headers)
    for j, h in enumerate(headers, start=1):
        col = get_column_letter(j)
        ws.column_dimensions[col].width = min(60, len(str(h)) + 2)
    ws.freeze_panes = f"A{header_row + 1}"
    last_row = header_row + max(n_data_rows, 1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(ncols)}{last_row}"
    )


def _write_excel_workbook(
    output_path: str,
    leads_rows: list[list[str]],
    dup_row_indices: set[int],
    errors: list[dict],
    dup_check_skipped: bool,
) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    ws_leads = wb.active
    ws_leads.title = "Лиды"
    ws_leads.append(BITRIX_EXCEL_HEADERS)
    for row in leads_rows:
        ws_leads.append(row)
    _apply_sheet_view(ws_leads, BITRIX_EXCEL_HEADERS, len(leads_rows), header_row=1)

    ws_dup = wb.create_sheet("Дубликаты")
    dup_rows_written = 0
    if dup_check_skipped:
        end_c = get_column_letter(len(BITRIX_EXCEL_HEADERS))
        ws_dup.merge_cells(f"A1:{end_c}1")
        ws_dup["A1"] = (
            f"⚠️ Строк в CSV больше {MAX_DUPLICATE_CHECK}, "
            "проверка дубликатов отключена"
        )
        hdr_row = 2
        for c, h in enumerate(BITRIX_EXCEL_HEADERS, start=1):
            ws_dup.cell(row=hdr_row, column=c, value=h)
        data_row = 3
        for i, row in enumerate(leads_rows):
            if i not in dup_row_indices:
                continue
            for c, val in enumerate(row, start=1):
                cell = ws_dup.cell(row=data_row, column=c, value=val)
                cell.fill = RED_FILL
            data_row += 1
            dup_rows_written += 1
        _apply_sheet_view(
            ws_dup, BITRIX_EXCEL_HEADERS, dup_rows_written, header_row=2
        )
    else:
        ws_dup.append(BITRIX_EXCEL_HEADERS)
        for i, row in enumerate(leads_rows):
            if i not in dup_row_indices:
                continue
            ws_dup.append(row)
            excel_row = ws_dup.max_row
            for c in range(1, len(BITRIX_EXCEL_HEADERS) + 1):
                ws_dup.cell(row=excel_row, column=c).fill = RED_FILL
            dup_rows_written += 1
        _apply_sheet_view(
            ws_dup, BITRIX_EXCEL_HEADERS, dup_rows_written, header_row=1
        )

    err_headers = list(BITRIX_EXCEL_HEADERS) + ["Ошибка"]
    ws_err = wb.create_sheet("Ошибки")
    ws_err.append(err_headers)
    for err in errors:
        raw = err.get("data") or {}
        line = _error_row_values(raw)
        line.append(str(err.get("error", "")))
        ws_err.append(line)
    for j, h in enumerate(err_headers, start=1):
        col = get_column_letter(j)
        ws_err.column_dimensions[col].width = min(60, len(str(h)) + 2)
    ws_err.freeze_panes = "A2"
    n_err = len(errors)
    last_err = 1 + max(n_err, 1)
    ws_err.auto_filter.ref = f"A1:{get_column_letter(len(err_headers))}{last_err}"

    wb.save(output_path)


def process_leads(
    input_path: str = INPUT_FILE,
    output_path: str = OUTPUT_EXCEL,
) -> dict:
    """Читает CSV, мапит, сохраняет Excel. Возвращает сводку для CLI."""
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Файл не найден: {input_path}")

    raw_leads = read_csv(input_path)
    mapped_leads: list[dict] = []
    errors: list[dict] = []
    leads_rows: list[list[str]] = []

    for i, raw in enumerate(raw_leads, start=1):
        if i % PROGRESS_EVERY == 0:
            print(f"… обработано строк: {i} / {len(raw_leads)}", flush=True)
        try:
            mapped = map_lead(raw)
            mapped["_csv_row"] = i
            mapped_leads.append(mapped)
            leads_rows.append(_build_excel_row(raw))
        except Exception as e:
            errors.append({"row": i, "error": str(e), "data": raw})

    n_rows = len(raw_leads)
    if n_rows > MAX_DUPLICATE_CHECK:
        print(
            f"⚠️ Строк в CSV больше {MAX_DUPLICATE_CHECK}, "
            "проверка дубликатов отключена",
            flush=True,
        )
        duplicates = {"phones": {}, "emails": {}, "companies": {}, "skipped": True}
        dup_count = 0
        dup_ids: set[str] = set()
        dup_check_skipped = True
    else:
        duplicates = _build_duplicates(mapped_leads)
        duplicates["skipped"] = False
        dup_count = _duplicate_groups_count(duplicates)
        dup_ids = _lead_ids_in_duplicates(duplicates)
        dup_check_skipped = False

    dup_row_indices: set[int] = set()
    for idx, mapped in enumerate(mapped_leads):
        lid = (mapped.get("id") or "").strip()
        if not lid:
            lid = f"row_{mapped.get('_csv_row', 0)}"
        if lid in dup_ids:
            dup_row_indices.add(idx)

    _write_excel_workbook(
        output_path, leads_rows, dup_row_indices, errors, dup_check_skipped
    )

    return {
        "processed": len(mapped_leads),
        "errors": len(errors),
        "duplicates_count": dup_count,
    }


if __name__ == "__main__":
    summary = process_leads()
    print(
        f"Обработано: {summary['processed']}, "
        f"ошибок: {summary['errors']}, "
        f"групп дубликатов: {summary['duplicates_count']}"
    )
